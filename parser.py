import asyncio
import csv
from datetime import datetime, date, time

import xlrd
from bs4 import BeautifulSoup
from openpyxl.reader.excel import load_workbook
from sqlalchemy.exc import IntegrityError

from database import FinamReportsDAO
from settings import logger


class Parser:
    """Класс парсинга файлов"""

    @staticmethod
    def import_type_detector(import_type: str) -> dict:
        """При добавлении импортов заполнить индексами столбцов начиная с нуля"""
        if import_type == 'finam':
            indexes = dict(
                import_type=import_type,
                date_index=0,
                date_time_index=1,
                type_index=2,
                comment_index=3,
                symbol_name_index=4,
                symbol_index=5,
                account_index=6,
                sum_index=7,
                date_format="%Y-%m-%d",
                date_time_format="%H:%M:%S"
            )
        else:
            indexes = dict()
        return indexes

    @staticmethod
    async def sql_update(update_dict: dict, cond_dict: dict, ignore_dict: dict, counter: int, import_type: str,
                         file: str):
        try:
            total_dict = {**cond_dict, **update_dict, **ignore_dict}
            await FinamReportsDAO.add(data=total_dict)
        except IntegrityError:
            try:
                await FinamReportsDAO.update(data=update_dict, conditions=cond_dict)
            except ValueError:
                logger.warning(f'File: {file} || {import_type} || Не удалось добавить позицию № {counter}. '
                               f'Неверный формат данных')
        except ValueError:
            logger.warning(f'File: {file} || {import_type} || Не удалось добавить позицию № {counter}. '
                           f'Неверный формат данных')

    @classmethod
    async def csv_parser(cls, file: str, indexes: dict, account_id: int):
        """Метод парсинга CSV"""

        with open(file, newline='') as csvfile:
            reader = csv.reader(csvfile)
            counter = 1
            for row in reader:
                data = {}
                if row[indexes['date_index']] != '':
                    data['date_record'] = datetime.strptime(row[indexes['date_index']], indexes['date_format'])
                else:
                    data['date_record'] = date(year=2001, month=1, day=1)

                if row[indexes['date_time_index']] != '':
                    if len(row[indexes['date_time_index']].split(':')) == 2:
                        date_time_index = f"{row[indexes['date_time_index']]}:00"
                    else:
                        date_time_index = row[indexes['date_time_index']]
                    data['date_time_record'] = datetime.strptime(date_time_index, indexes['date_time_format']).time()
                else:
                    data['date_time_record'] = time(hour=0, minute=0, second=0)

                data['type_record'] = row[indexes['type_index']]
                data['comment_record'] = row[indexes['comment_index']]
                data['symbol_name_record'] = row[indexes['symbol_name_index']]
                data['symbol_record'] = row[indexes['symbol_index']]
                data['account_record'] = row[indexes['account_index']]
                data['account_id_record'] = account_id
                data['sum_record'] = float(row[indexes['sum_index']]) if row[indexes['sum_index']] != '' else 0
                try:
                    await FinamReportsDAO.add(data)
                except IntegrityError:
                    logger.warning(
                        f'File: {file} || {indexes["import_type"]} || Не удалось добавить позицию № {counter}. Строка '
                        f'уже существует')
                except ValueError:
                    logger.warning(
                        f'File: {file} || {indexes["import_type"]} || Не удалось добавить позицию № {counter}. '
                        f'Неверный формат данных')
                counter += 1

    @classmethod
    async def xml_parser(cls, file: str, account_id: int, import_type: str):
        """Метод парсинга XML"""
        with open(file, 'r') as file:
            xml_file = file.read()
            soup = BeautifulSoup(xml_file, 'lxml')

            operation_block = soup.find('db9').find_all('r')
            counter = 1
            for operation in operation_block:
                update_dict = {}  # Словарь для новых данных
                cond_dict = {}  # Словарь для выполнения условий
                ignore_dict = {}  # Словарь для неизменяемых данных
                if operation.get('d') != '':
                    cond_dict['date_record'] = datetime.strptime(operation.get('d'), "%d.%m.%Y")
                else:
                    cond_dict['date_record'] = date(year=2001, month=1, day=1)
                if operation.get('t') != '':
                    cond_dict['date_time_record'] = datetime.strptime(operation.get('t'), "%H:%M:%S").time()
                else:
                    cond_dict['date_time_record'] = time(hour=0, minute=0, second=0)
                ignore_dict['type_record'] = operation.get('op')
                ignore_dict['comment_record'] = operation.get('c')
                cond_dict['symbol_name_record'] = operation.get('is')
                update_dict['isin_record'] = operation.get('isin')
                update_dict['count_record'] = int(float(operation.get('qty'))) if operation.get('qty') is not None else 0
                update_dict['deal_price_record'] = float(operation.get('pr')) if operation.get('pr') is not None else 0
                cond_dict['sum_record'] = float(operation.get('spra')) if operation.get('spra') is not None else 0
                update_dict['deal_id_record'] = operation.get('trdn')
                update_dict['account_prefix_record'] = soup.find('account').get('name').split('-')[0]
                cond_dict['account_record'] = soup.find('account').get('id')
                cond_dict['account_id_record'] = account_id  # Уточнить является ли условием

                await cls.sql_update(
                    update_dict=update_dict,
                    cond_dict=cond_dict,
                    ignore_dict=ignore_dict,
                    file=file,
                    counter=counter,
                    import_type=import_type
                )
                counter += 1

    @classmethod
    async def xls_parser(cls, file: str, account_id: int, import_type: str):
        """Метод парсинга xls"""
        with xlrd.open_workbook(filename=file) as wb:
            sh = wb.sheet_by_index(0)
            counter = 1
            for rx in range(sh.nrows)[49:-4]:
                update_dict = {}  # Словарь для новых данных
                cond_dict = {}  # Словарь для выполнения условий
                ignore_dict = {}  # Словарь для неизменяемых данных
                cond_dict['date_record'] = xlrd.xldate_as_datetime(sh[rx][0].value, 0).date()
                cond_dict['date_time_record'] = xlrd.xldate_as_datetime(sh[rx][1].value, 0).time()
                cond_dict['symbol_name_record'] = sh[rx][2].value
                update_dict['account_prefix_record'] = sh[4][1].value.split('-')[0]
                cond_dict['account_record'] = sh[4][1].value.split('-')[1]
                cond_dict['account_id_record'] = account_id
                update_dict['isin_record'] = sh[rx][3].value
                ignore_dict['type_record'] = sh[rx][4].value
                update_dict['count_record'] = int(sh[rx][5].value) if sh[rx][5].value != '' else 0
                update_dict['deal_price_record'] = float(sh[rx][6].value) if sh[rx][6].value != '' else 0
                cond_dict['sum_record'] = float(sh[rx][13].value) if sh[rx][10].value != '' else 0
                update_dict['deal_id_record'] = sh[rx][25].value
                ignore_dict['comment_record'] = sh[rx][27].value
                await cls.sql_update(
                    update_dict=update_dict,
                    cond_dict=cond_dict,
                    ignore_dict=ignore_dict,
                    file=file,
                    counter=counter,
                    import_type=import_type
                )
                counter += 1

    @classmethod
    async def xlsx_parser(cls, file: str, account_id: int, import_type: str):
        """Метод парсинга xlsx"""
        wb = load_workbook(filename=file)
        sh = wb.active
        counter = 1
        for row in sh.iter_rows(min_row=50, max_row=sh.max_row-4):
            update_dict = {}  # Словарь для новых данных
            cond_dict = {}  # Словарь для выполнения условий
            ignore_dict = {}  # Словарь для неизменяемых данных
            cond_dict['date_record'] = row[0].value.date()
            cond_dict['date_time_record'] = row[1].value
            cond_dict['symbol_name_record'] = row[2].value
            update_dict['account_prefix_record'] = sh['B5'].value.split('-')[0]
            cond_dict['account_record'] = sh['B5'].value.split('-')[1]
            cond_dict['account_id_record'] = account_id
            update_dict['isin_record'] = row[3].value
            ignore_dict['type_record'] = row[4].value
            update_dict['count_record'] = int(row[5].value) if row[5].value is not None else 0
            update_dict['deal_price_record'] = float(row[6].value) if row[6].value is not None else 0
            cond_dict['sum_record'] = float(row[13].value) if row[13].value is not None else 0
            update_dict['deal_id_record'] = row[25].value
            ignore_dict['comment_record'] = row[27].value
            await cls.sql_update(
                update_dict=update_dict,
                cond_dict=cond_dict,
                ignore_dict=ignore_dict,
                file=file,
                counter=counter,
                import_type=import_type
            )
            counter += 1
